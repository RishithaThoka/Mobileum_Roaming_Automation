import openpyxl
from openpyxl.styles import Font

def build(path, values):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    for sheet_name, rows in values.items():
        ws = wb.create_sheet(sheet_name)
        ws.append(["Field", "Value"])
        for cell in ws[1]:
            cell.font = Font(bold=True, name="Arial")
        for label, value in rows:
            ws.append([label, value])
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = Font(name="Arial")
        ws.column_dimensions['A'].width = 34
        ws.column_dimensions['B'].width = 30
    wb.save(path)

v1 = {
    "NetworkTechnical": [
        ("HLR Global Title", "34916230000"),
        ("Signalling Point Code", "2-45-6"),
        ("GRX IPX Provider", "Tata Communications"),
        ("APN Roaming", "telefonica.es.roaming"),
    ],
    "Security": [
        ("Steering Of Roaming Supported", "true"),
        ("Fraud Contact Email", "fraud@telefonica-test.example"),
    ],
    "Commercial": [
        ("Roaming Agreement Type", "Bilateral"),
        ("Preferred Roaming Partner Tier", "Tier2"),
    ],
    "FinancialBilling": [
        ("Wholesale Voice Rate", "0.0165"),
        ("Wholesale Data Rate Per MB", "0.0052"),
        ("Currency", "EUR"),
        ("Settlement Contact", "settlements@telefonica-test.example"),
    ],
    "Operations": [
        ("Roaming Operations Manager", "Marta Ruiz"),
        ("Support Hours", "24x7"),
        ("Effective Date", "2026-03-01"),
    ],
}

v2 = {
    "NetworkTechnical": [
        ("HLR Global Title", "34916239999"),
        ("Signalling Point Code", "2-45-6"),
        ("GRX IPX Provider", "Tata Communications"),
        ("APN Roaming", "telefonica.es.roaming2"),
    ],
    "Security": [
        ("Steering Of Roaming Supported", "true"),
        ("Fraud Contact Email", "fraud@telefonica-test.example"),
    ],
    "Commercial": [
        ("Roaming Agreement Type", "Bilateral"),
        ("Preferred Roaming Partner Tier", "Tier1"),
    ],
    "FinancialBilling": [
        ("Wholesale Voice Rate", "0.0190"),
        ("Wholesale Data Rate Per MB", "0.0041"),
        ("Currency", "EUR"),
        ("Settlement Contact", "settlements@telefonica-test.example"),
    ],
    "Operations": [
        ("Roaming Operations Manager", "Marta Ruiz"),
        ("Support Hours", "24x7"),
        ("Effective Date", "2026-09-01"),
    ],
}

build("TelefonicaEspana_IR21_v1.xlsx", v1)
build("TelefonicaEspana_IR21_v2.xlsx", v2)
print("done")
